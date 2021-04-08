import os, sys, re, json, collections, time, enum, shutil
from operator import itemgetter, attrgetter, methodcaller
from collections import Counter, OrderedDict
from dataclasses import dataclass, field
from datetime import datetime
import pathlib, asyncio, copy
import openpyxl

from src.model_product import *
from src.module_xml import *

class ReportFile:
  def __init__(self):
    self.path = None
    self.wb = None
    # after complete report status set to true
    self.status = False

  def preview(self):
    # preview output file after operation complete
    status = False
    if (self.status == True):
      # check if valid file
      if (self.path != None):
        os.system("start excel ""{0}""".format(self.path))
        status = True
    return status

  def load_wb(self):
    # if report file exist
    if (os.path.exists(self.path)):
      # load workbook
      self.wb = openpyxl.load_workbook(self.path)
      return True
    else:
      return False

  def save_wb(self, _path=None):
    try:
      # if workbook exist
      if (self.wb != None):
        # check if target _path existed
        if (_path == None):
          # save to previous imported file
          self.wb.save(self.path)
          self.wb.close()
        else:
          if (os.path.exists(_path)):
              # remove old one
              os.remove(_path)
          # check if dir existed
          elif (not os.path.exists(os.path.dirname(_path))):
            # make leaf dir and intermediate one
            os.makedirs(os.path.dirname(_path))
          else:
            pass
          self.wb.save(_path)
          self.wb.close()
        return True
      else:
        return False
    except:
      return False

class ReportApi:
  def __init__(self, _console=None, _parent=None):
    self.parent = _parent
    self.amazon_prd_list = []
    self.ebay_prd_list = []
    self.console = _console
    self.profit_formula = None
    self.profit_col = None
    self.alert_col = None
    self.convert_col = None
    self.img_col = None
    self.img_link_col = None
    self.data_dict = None
    self.__excel_location = None
    self.__start_row = 0
    self.__col_idx = 0
    self.__col = {"AMAZON-LOC":{},
                  "EBAY-LOC":{}}
    self.__report_file = ReportFile()
    self.__loading_excel_ele_location()

  def __loading_excel_ele_location(self):
    # load data from excel_element_location.xml file
    self.__excel_location = XmlDoc("excel_element_location.xml", _console=self.console)
    self.__excel_location.parse()
    self.data_dict = self.__excel_location.get_dict()
    self.__start_row, self.__col_idx = openpyxl.utils.coordinate_to_tuple(self.data_dict["INDEX-COL"])
    self.__start_row += 1
    try:
      for loc in ["AMAZON-LOC", "EBAY-LOC"]:
        coord = self.data_dict[loc]
        _, self.__col[loc]["LINK"] = openpyxl.utils.coordinate_to_tuple(coord["LINK"])
        _, self.__col[loc]["ITEM-TITLE"]  = openpyxl.utils.coordinate_to_tuple(coord["ITEM-TITLE"])
        _, self.__col[loc]["STATUS"]  = openpyxl.utils.coordinate_to_tuple(coord["STATUS"])
        _, self.__col[loc]["PRICE"]  = openpyxl.utils.coordinate_to_tuple(coord["PRICE"])
      _, self.profit_col = openpyxl.utils.coordinate_to_tuple(self.data_dict["PROFIT"])
      _, self.alert_col = openpyxl.utils.coordinate_to_tuple(self.data_dict["ALERT"])
      _, self.convert_col = openpyxl.utils.coordinate_to_tuple(self.data_dict["CONVERT"])
      _, self.img_col = openpyxl.utils.coordinate_to_tuple(self.data_dict["IMG"])
      _, self.img_link_col = openpyxl.utils.coordinate_to_tuple(self.data_dict["IMG-LINK"])
    except Exception as e:
      self.__console_log(f"Error occur during get data from excel_element_location dict, {str(e)}")


  def __console_log(self, val:str):
    if (self.console != None):
      self.console(f"[REPORT]: {val}")

  def set_report_file(self, file_path):
    # import report file
    # set report file path
    if (os.path.isfile(file_path)):
      _, extension = os.path.splitext(file_path)
      if (os.path.exists(os.path.abspath(file_path)) and (extension == ".xlsx")):
        self.__report_file.path = os.path.abspath(file_path)
        return True
      else:
        return False
    else:
      return False

  def preview(self):
    if(False == self.__report_file.preview()):
      self.__console_log("Report file is not completed")

  def save_report(self, file_path=None):
    # save wb to selected file
    return self.__report_file.save_wb(file_path)

  def get_prd_link(self):
    try:
      if (self.__report_file.load_wb()):
        ws = self.__report_file.wb.active
        # clear amazon, ebay product list and update new data
        self.amazon_prd_list.clear()
        self.ebay_prd_list.clear()
        for row in range(self.__start_row, ws.max_row+1):
          for loc in ["AMAZON-LOC", "EBAY-LOC"]:
            link_val = ws.cell(row, self.__col[loc]["LINK"]).value
            item_number = ws.cell(row, self.__col_idx).value
            if ((link_val != None) and (link_val != "")):
              if ((item_number != None) and (item_number != "")):
                temp_prod = Product(_link=link_val,_row_idx=row, _item_num=int(item_number))
              else:
                temp_prod = Product(_link=link_val,_row_idx=row, _item_num=1)
                self.__console_log(f"Item at row {row} missing link or SI number")
              if (loc == "AMAZON-LOC"):
                self.amazon_prd_list.append(temp_prod)
              else:
                self.ebay_prd_list.append(temp_prod)
        return True
      else:
        return False
    except Exception as e:
      self.__console_log(f"Error occur during get product information from excel file. {str(e)}")
      return False

  def __get_val_formula(self, row, dict_val):
    new_coord = ["{0}{1}".format(i, row) for i in dict_val["col"]]
    dict_val["val_list"][::2] = dict_val["base"]
    dict_val["val_list"][1::2] = new_coord
    formula_val = "".join(dict_val["val_list"])
    return formula_val

  def gen_report(self):
    try:
      ws = self.__report_file.wb.active
      # set profit formula self.profit_formula
      if (self.profit_formula == "" or self.profit_formula == None):
        self.profit_formula = self.data_dict["DEFAULT-VAL"]["FORMULA-VAL"]

      output_dict = {}
      for idx, str_val in enumerate([self.profit_formula, self.data_dict["DEFAULT-VAL"]["ALERT-VAL"], self.data_dict["DEFAULT-VAL"]["CONVERT-VAL"]]):
        base_str = re.split(r"\$?[A-Z]{1,3}\$?\d+", str_val)
        col_list = []
        for item in re.finditer(r"\$?([A-Z]{1,3})\$?\d+", str_val):
          col_list.append(item.group(1))
        output_dict[idx] = {}
        output_dict[idx]["base"] = base_str
        output_dict[idx]["col"] = col_list
        output_dict[idx]["val_list"] = [None]*(len(base_str) + len(col_list))

      style_list = []
      style_list.append(copy.copy(ws.cell(self.amazon_prd_list[0].row_idx, self.__col["AMAZON-LOC"]["PRICE"])._style))
      style_list.append(copy.copy(ws.cell(self.amazon_prd_list[0].row_idx, self.__col["AMAZON-LOC"]["ITEM-TITLE"])._style))
      style_list.append(copy.copy(ws.cell(self.amazon_prd_list[0].row_idx, self.__col["AMAZON-LOC"]["STATUS"])._style))
      style_list.append(copy.copy(ws.cell(self.amazon_prd_list[0].row_idx, self.profit_col)._style))
      style_list.append(copy.copy(ws.cell(self.amazon_prd_list[0].row_idx, self.alert_col)._style))
      style_list.append(copy.copy(ws.cell(self.amazon_prd_list[0].row_idx, self.convert_col)._style))
      style_list.append(copy.copy(ws.cell(self.amazon_prd_list[0].row_idx, self.__col["EBAY-LOC"]["PRICE"])._style))
      style_list.append(copy.copy(ws.cell(self.amazon_prd_list[0].row_idx, self.__col["EBAY-LOC"]["ITEM-TITLE"])._style))
      style_list.append(copy.copy(ws.cell(self.amazon_prd_list[0].row_idx, self.__col["EBAY-LOC"]["STATUS"])._style))

      for item in self.amazon_prd_list:
        formula_val = self.__get_val_formula(item.row_idx, output_dict[0])
        alert_val = self.__get_val_formula(item.row_idx, output_dict[1])
        convert_val = self.__get_val_formula(item.row_idx, output_dict[2])

        ws.cell(item.row_idx, self.__col["AMAZON-LOC"]["PRICE"]).value = item.price
        ws.cell(item.row_idx, self.__col["AMAZON-LOC"]["ITEM-TITLE"]).value = item.name
        ws.cell(item.row_idx, self.__col["AMAZON-LOC"]["STATUS"]).value = ITEM_STATUS[item.status]
        # update profit formula
        ws.cell(item.row_idx, self.profit_col).value = str(formula_val)
        # update alert
        ws.cell(item.row_idx, self.alert_col).value = str(alert_val)
        # update convert
        ws.cell(item.row_idx, self.convert_col).value = str(convert_val)
        # update img
        # ws.cell(item.row_idx, self.img_col)
        # update img link
        # ws.cell(item.row_idx, self.img_link_col)
        ## update style
        ws.cell(item.row_idx, self.__col["AMAZON-LOC"]["PRICE"])._style = style_list[0]
        ws.cell(item.row_idx, self.__col["AMAZON-LOC"]["ITEM-TITLE"])._style = style_list[1]
        ws.cell(item.row_idx, self.__col["AMAZON-LOC"]["STATUS"])._style = style_list[2]
        ws.cell(item.row_idx, self.profit_col)._style = style_list[3]
        ws.cell(item.row_idx, self.alert_col)._style = style_list[4]
        ws.cell(item.row_idx, self.convert_col)._style = style_list[5]

      for item in self.ebay_prd_list:
        ws.cell(item.row_idx, self.__col["EBAY-LOC"]["PRICE"]).value = item.price
        ws.cell(item.row_idx, self.__col["EBAY-LOC"]["ITEM-TITLE"]).value = item.name
        ws.cell(item.row_idx, self.__col["EBAY-LOC"]["STATUS"]).value = ITEM_STATUS[item.status]

        ws.cell(item.row_idx, self.__col["EBAY-LOC"]["PRICE"])._style = style_list[6]
        ws.cell(item.row_idx, self.__col["EBAY-LOC"]["ITEM-TITLE"])._style = style_list[7]
        ws.cell(item.row_idx, self.__col["EBAY-LOC"]["STATUS"])._style = style_list[8]
      self.__report_file.save_wb()
      self.__report_file.status = True
      return True
    except Exception as e:
      self.__console_log(f"Error occur during generate report file, {str(e)}")
      return False

  # giá ebay - (giá ebay nhân 15% + giá amz) <0 thì gửi email kèm hìghlight trên file excel