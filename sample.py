import os, sys, re, json, collections, time
from operator import itemgetter, attrgetter, methodcaller
from collections import Counter, OrderedDict
from dataclasses import dataclass, field
from datetime import datetime
import pathlib, asyncio
import openpyxl
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as wait
COL_EBAY_LINK = 2
COL_ITEM_TITLE = 4
COL_AMA_LINK = 5
COL_AMA_STS = 6
COL_AMA_PRICE = 8
COL_EBAY_PRICE = 7

class ItemStatus:
  NONE = 1
  IN_STOCK = 2
  OUT_STOCK = 3

@dataclass
class Product:
  site:int
  link:str
  name:str
  price:float
  currency:str
  status:ItemStatus
  multi_vendor:bool
  def __init__(self, _link="", _name="", _price=0, _cur="", _status=ItemStatus.NONE, _mul=False, _site=0):
    self.link = _link
    self.name = _name
    self.price = _price
    self.currency = _cur
    self.status = _status
    self.multi_vendor = _mul
    self.site = _site

def get_price_ebay(prod_list):

  pass

def click_element(driver, xpath):
  try:
    wait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
    return True
  except:
    return False
def send_data(driver, xpath, data):
  try:
    wait(driver, 5).until(EC.presence_of_element_located((By.XPATH,xpath))).send_keys(data)
    return True
  except:
    return False

def get_price_amazon(prod_list):
  # TODO: can get data faster with request(grequest), but has to change the zipcode to correct location with the request?
  #       issue: ip block with multiple request
  # header = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.190 Safari/537.36"}
  # rs = [requests.get(item.link, headers=header) for item in prod_list]
  # for idx,response in enumerate(rs):
  #   beu = BeautifulSoup(response.content, "html.parser")
  #   with open("html_item{0}.html".format(idx),"w",encoding="utf-8") as f:
  #     f.write(beu.decode())
  # pass

  # SELENIUM
  options = webdriver.ChromeOptions()
  options.add_argument("--ignore-certificate-errors")
  options.add_argument("--incognito")
  # options.add_argument("--headless")
  driver = webdriver.Chrome("./browserdriver/ChromeDriver88.0.4324.96_32b.exe", chrome_options=options)
  driver.get("https://amazon.com")

  # find zipcode element
  click_element(driver, '//*[@id="nav-global-location-popover-link"]')
  # //*[@id="GLUXZipUpdateInput"]
  send_data(driver, '//*[@id="GLUXZipUpdateInput"]', "97124")
  # GLUXZipUpdate - update code button
  click_element(driver, '//*[@id="GLUXZipUpdate"]')
  # continue - GLUXConfirmAction - /html/body/div[5]/div/div/div[2]/span/span
  if (False == click_element(driver,'//*/div[@class="a-popover-wrapper"]/div[@class="a-popover-footer"]/span/span')):
    # a-autoid-29-announce - apply code
    click_element(driver,'//*[@id="a-autoid-29-announce"]')
  for item in prod_list:
    driver.get(item.link)

  driver.quit()

def update_excel(prod_list):

  pass

def main(args)->None:
  wb = openpyxl.load_workbook(args[1])
  ws = wb["Sheet1"]

  prod_list = {0:[], 1:[]}
  for row in range(2, ws.max_row):
    for idx, col in enumerate([COL_EBAY_LINK, COL_AMA_LINK]):
      val = ws.cell(row, col).value
      if ((val != None) and (val != "")):
        temp_prod = Product(_site=col,_link=val)
        prod_list[idx].append(temp_prod)

  get_price_ebay(prod_list[0])

  get_price_amazon(prod_list[1])

  update_excel(prod_list)

  pass

if __name__ == "__main__":
  main(sys.argv)