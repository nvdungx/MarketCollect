import os, sys, re, json, collections, time
from operator import itemgetter, attrgetter, methodcaller
from collections import Counter, OrderedDict
from dataclasses import dataclass, field
from datetime import datetime
import pathlib, asyncio
import openpyxl
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as wait
COL_EBAY_LINK = 2
COL_ITEM_TITLE = 4
COL_AMA_LINK = 5
COL_AMA_STS = 6
COL_AMA_PRICE = 8
COL_EBAY_PRICE = 7

class ItemStatus:
  NONE = 1
  IN_STOCK = 2
  OUT_STOCK = 3

ITEM_STATUS = {
  ItemStatus.NONE:"NONE",
  ItemStatus.IN_STOCK:"In Stock",
  ItemStatus.OUT_STOCK:"Out of Stock"
}

@dataclass
class Product:
  site:int
  row_idx:int
  link:str
  name:str
  price:float
  currency:str
  status:ItemStatus
  status_str:str
  multi_vendor:bool
  def __init__(self, _link="", _name="", _price=0, _cur="", _status=ItemStatus.NONE, _mul=False, _site=0, _row_idx=1):
    # product link
    self.link = _link
    # product name
    self.name = _name
    # price value
    self.price = _price
    # currency
    self.currency = _cur
    # ItemStatus
    self.status = _status
    self.status_str = ""
    # if product is provided by multiple vendor
    self.multi_vendor = _mul
    # 5[ama] 2[ebay]
    self.site = _site
    self.row_idx = _row_idx

def get_price_ebay(prod_list):

  pass

def click_element(driver, xpath):
  try:
    wait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,xpath))).click()
    return True
  except:
    return False

def send_data(driver, xpath, data):
  try:
    wait(driver, 5).until(EC.presence_of_element_located((By.XPATH,xpath))).send_keys(data)
    return True
  except:
    return False

def get_element(driver, xpath):
  try:
    ele = driver.find_element_by_xpath(xpath)
    return True, ele
  except:
    return False, None

def get_price_amazon(prod_list):
  # GENERAL ISSUE WITH THESE WEB crawl APPROACH: HTML structure of page could be change(when amazon change there page rendering)
  # this lead to tool need to be re update with new xpath value for each element(name, price, etc..)

  # TODO: can get data faster with request(grequest), but has to change the zipcode to correct location with the request?
  #       issue: ip block with multiple request
  # header = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.190 Safari/537.36"}
  # rs = [requests.get(item.link, headers=header) for item in prod_list]
  # for idx,response in enumerate(rs):
  #   beu = BeautifulSoup(response.content, "html.parser")
  #   with open("html_item{0}.html".format(idx),"w",encoding="utf-8") as f:
  #     f.write(beu.decode())
  # pass

  # SELENIUM
  options = webdriver.ChromeOptions()
  options.add_argument("--ignore-certificate-errors")
  options.add_argument("--incognito")
  # options.add_argument("--headless")
  driver = webdriver.Chrome("./browserdriver/ChromeDriver88.0.4324.96_32b.exe", chrome_options=options)
  driver.get("https://amazon.com")
  # find zipcode element
  if(not click_element(driver, '//*[@id="nav-global-location-popover-link"]')):
    raise Exception("WEB_DRIVER", "Failed to find amazon html element {0}".format("zip-code-button"))
  # //*[@id="GLUXZipUpdateInput"]
  if(not send_data(driver, '//*[@id="GLUXZipUpdateInput"]', "97124")):
    raise Exception("WEB_DRIVER", "Failed to find amazon html element {0}".format("zip-code-textbox"))
  # GLUXZipUpdate - update code button
  if(not click_element(driver, '//*[@id="GLUXZipUpdate"]')):
    raise Exception("WEB_DRIVER", "Failed to find amazon html element {0}".format("zip-code-update-button"))
  # continue - GLUXConfirmAction - /html/body/div[5]/div/div/div[2]/span/span
  if (False == click_element(driver,'//*/div[@class="a-popover-wrapper"]/div[@class="a-popover-footer"]/span/span')):
    # a-autoid-29-announce - apply code
    if(not click_element(driver,'//*[@id="a-autoid-29-announce"]')):
      raise Exception("WEB_DRIVER", "Failed to find amazon html element {0}".format("zip-code-continue-button"))

  for item_idx, item in enumerate(prod_list):
    driver.get(item.link)
    time.sleep(1)
    # TITLE
    # //*/div[@id="centerCol"]/*/[@id="productTitle"]
    result, ele = get_element(driver, '//*/span[@id="productTitle"]')
    if(not result):
      raise Exception("WEB_DRIVER", "Failed to find amazon html element {0}".format("product-title"))
    else:
      item.name = ele.text.strip()
    # PRICE
    #             id="a-page"/id="dp"/id="dp-container"/id="ppd"/id="centerCol"/id="desktop_unifiedPrice"(div[10])/id="unifiedPrice_feature_div"/id="price"/class="a-lineitem"(table)
    #                                                 id="priceblock_ourprice_row"(tr)/class="a-span12"(td[2])/id="priceblock_ourprice"
    # /html/body/div[2]/div[3]/div[7]/div[5]/div[4]/div[10]/div/div/table/tbody/tr/td[2]/span[1]
    # //*[@id="priceblock_ourprice"]
    result, ele = get_element(driver, '//*/span[@id="priceblock_ourprice"]')
    if(not result):
      result, ele = get_element(driver, '//*/span[@id="priceblock_saleprice"]')
      if (not result):
        pass
        # print("Failed to find amazon html element {0}".format("product-price"))
    if(result):
      item.multi_vendor = False
      item.status = ItemStatus.IN_STOCK
      temp = ele.text.strip()
      reg = re.match(r"([^\d\.]*)([\d\.]+)([^\d\.]*)",temp)
      if (reg != None):
        item.price = float(str(reg[2]))
        if(str(reg[1]).strip() != ""):
          item.currency = str(reg[1]).strip()
        elif(str(reg[3]).strip() != ""):
          item.currency = str(reg[1]).strip()
        else:
          print("Cannot find product currency")
      else:
        raise Exception("PROCESS", "Failed to get price from string {0}".format(temp))

    # ITEM OUT OF STOCK or AMAZON PLACE "Available from these sellers." span script
    if (not result):
      # /html/body/div[2]/div[2]/div[8]/div[4]/div[4]/div[19]/div[1]/span
      # //*[@id="availability"]/span
      result, ele = get_element(driver, '//*[@id="availability"]/span')
      if (not result):
        print("Failed to find amazon html element {0}".format("product-availability"))
      else:
        # OUT OF STOCK
        if (ele.text.strip() == ""):
          item.status_str = ele.text.strip()
          item.status = ItemStatus.OUT_STOCK
          item.price = 0
          item.currency = "$"
          print("Item {0} is out of stock".format(item_idx))
        # Available from these sellers.
        else:
          # get price from vender offer option
          result = False
    # ITEM HAS MULTIPLE VENDOR
    # /html/body/div[2]/div[2]/div[9]/div[4]/div[4]/div[44]
    # /html/body/div[2]/div[2]/div[8]/div[4]/div[4]/div[44]/div[2]/span/a
    # //*[@id="olp_feature_div"]
    if (not result):
      result, ele = get_element(driver, '//*[@id="olp_feature_div"]/div/span/a[@class="a-link-normal"]/span[@class="a-size-base a-color-price"]')
      if (not result):
        raise Exception("WEB_DRIVER", "Failed to find product price, please update the target element xpath")
      else:
        item.multi_vendor = True
        item.status = ItemStatus.IN_STOCK
        temp = ele.text.strip()
        reg = re.match(r"([^\d\.]*)([\d\.]+)([^\d\.]*)",temp)
        if (reg != None):
          item.price = float(str(reg[2]))
          if(str(reg[1]).strip() != ""):
            item.currency = str(reg[1]).strip()
          elif(str(reg[3]).strip() != ""):
            item.currency = str(reg[1]).strip()
          else:
            print("Cannot find product currency")
        else:
          raise Exception("PROCESS", "Failed to get price from string {0}".format(temp))
    if (result):
      print("COLLECT ITEM {0} DONE - {1}{2}".format(item_idx, item.price, item.currency))
  driver.quit()

def update_excel(ws, prod_list):
  for item in prod_list[1]:
    ws.cell(item.row_idx, COL_AMA_PRICE).value = item.price
    ws.cell(item.row_idx, COL_ITEM_TITLE).value = item.name
    ws.cell(item.row_idx, COL_AMA_STS).value = ITEM_STATUS[item.status]
  print("--- DONE ----")

def main(args)->None:
  wb = openpyxl.load_workbook(args[1])
  ws = wb["Sheet1"]

  prod_list = {0:[], 1:[]}
  for row in range(2, ws.max_row):
    for idx, col in enumerate([COL_EBAY_LINK, COL_AMA_LINK]):
      val = ws.cell(row, col).value
      if ((val != None) and (val != "")):
        temp_prod = Product(_site=col,_link=val,_row_idx=row)
        prod_list[idx].append(temp_prod)

  get_price_ebay(prod_list[0])

  get_price_amazon(prod_list[1])

  update_excel(ws, prod_list)
  wb.save(args[1])
  wb.close()
  pass

if __name__ == "__main__":
  main(sys.argv)